"""
Generate Demo Excel Workbook for Hub Scoring Walkthrough
=========================================================
Creates a multi-sheet Excel file matching docs/DEMO_SCORING_WALKTHROUGH.md
with formatted tables, formulas, and color coding for presentation.
"""

import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Output path
# ---------------------------------------------------------------------------
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "results"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH = OUTPUT_DIR / "demo_scoring_walkthrough.xlsx"

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
NAVY = "1F3864"
DARK_BLUE = "2E75B6"
MEDIUM_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
VERY_LIGHT_BLUE = "E9EFF7"
GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "FCE4EC"
GOLD = "BF8F00"
LIGHT_GOLD = "FFF2CC"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
DARK_GRAY = "404040"

# ---------------------------------------------------------------------------
# Reusable style helpers
# ---------------------------------------------------------------------------
thin_border = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

thick_bottom = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="medium", color="404040"),
)


def _header_fill(hex_color=NAVY):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _row_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_font_dark = Font(name="Calibri", bold=True, color=DARK_GRAY, size=11)
title_font = Font(name="Calibri", bold=True, color=NAVY, size=16)
subtitle_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=13)
section_font = Font(name="Calibri", bold=True, color=MEDIUM_BLUE, size=11)
body_font = Font(name="Calibri", size=11, color=DARK_GRAY)
bold_font = Font(name="Calibri", bold=True, size=11, color=DARK_GRAY)
score_font = Font(name="Calibri", bold=True, size=12, color=NAVY)
number_fmt_0 = "#,##0"
number_fmt_2 = "#,##0.00"
number_fmt_3 = "#,##0.000"
pct_fmt = "0.0%"

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")


def write_header_row(ws, row, headers, start_col=1, fill_color=NAVY):
    """Write a formatted header row."""
    for c, text in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = header_font
        cell.fill = _header_fill(fill_color)
        cell.alignment = center_align
        cell.border = thin_border


def write_data_row(ws, row, values, start_col=1, fill_color=None, bold_cols=None,
                   num_fmt=None, font_override=None):
    """Write a formatted data row."""
    bold_cols = bold_cols or []
    for c, val in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=c, value=val)
        if c in bold_cols:
            cell.font = bold_font if font_override is None else font_override
        else:
            cell.font = body_font
        if fill_color:
            cell.fill = _row_fill(fill_color)
        cell.alignment = center_align
        cell.border = thin_border
        if num_fmt and isinstance(val, (int, float)):
            cell.number_format = num_fmt


def write_title(ws, row, col, text, font=title_font):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font
    cell.alignment = Alignment(horizontal="left", vertical="center")


def auto_width(ws, min_width=10, max_width=28):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = best


def set_col_widths(ws, widths):
    """Set explicit column widths from a dict {col_number: width}."""
    for col_num, w in widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = w


# ---------------------------------------------------------------------------
# Tier colours for row-level formatting
# ---------------------------------------------------------------------------
TIER_ROW_COLORS = {
    "national": LIGHT_BLUE,
    "metro": LIGHT_ORANGE,
}


def tier_fill(tier_key):
    return TIER_ROW_COLORS.get(tier_key, None)


# ===========================================================================
# Hub data (matching the markdown walkthrough exactly)
# ===========================================================================
HUBS = [
    {
        "id": 1, "name": "Tel Aviv Savidor", "hebrew": 'ת"א סבידור',
        "tier": "ארצי", "tier_en": "National", "tier_key": "national",
        "area": "תל אביב", "area_en": "Tel Aviv",
        "position": "גלעין", "position_en": "Core",
        "passengers": 150_000,
        "modes": ["Rail", "Metro", "LRT", "BRT"], "total_lines": 12,
        "pop": [45_000, 35_000, 25_000], "emp": [120_000, 80_000, 40_000],
        "near_terminal": True, "terminal_type": "מתקן משולב (Integrated)",
        "terminal_weight": 3.0,
    },
    {
        "id": 2, "name": "Jerusalem Central", "hebrew": "ירושלים מרכז",
        "tier": "ארצי", "tier_en": "National", "tier_key": "national",
        "area": "ירושלים", "area_en": "Jerusalem",
        "position": "גלעין", "position_en": "Core",
        "passengers": 85_000,
        "modes": ["Rail", "LRT", "BRT"], "total_lines": 8,
        "pop": [55_000, 42_000, 30_000], "emp": [70_000, 45_000, 25_000],
        "near_terminal": True, "terminal_type": "מסוף גדול (Large)",
        "terminal_weight": 3.0,
    },
    {
        "id": 3, "name": "Haifa Merkaz", "hebrew": "חיפה מרכז",
        "tier": "ארצי", "tier_en": "National", "tier_key": "national",
        "area": "חיפה", "area_en": "Haifa",
        "position": "גלעין", "position_en": "Core",
        "passengers": 60_000,
        "modes": ["Rail", "Metro", "BRT"], "total_lines": 7,
        "pop": [38_000, 28_000, 20_000], "emp": [55_000, 35_000, 18_000],
        "near_terminal": True, "terminal_type": "מסוף גדול (Large)",
        "terminal_weight": 3.0,
    },
    {
        "id": 4, "name": "Petah Tikva", "hebrew": "פתח תקווה",
        "tier": "מטרופוליני", "tier_en": "Metropolitan", "tier_key": "metro",
        "area": "תל אביב", "area_en": "Tel Aviv",
        "position": "טבעת פנימית", "position_en": "Inner Ring",
        "passengers": 32_000,
        "modes": ["Metro", "LRT"], "total_lines": 5,
        "pop": [60_000, 48_000, 35_000], "emp": [40_000, 25_000, 12_000],
        "near_terminal": True, "terminal_type": "מסוף בינוני (Medium)",
        "terminal_weight": 2.0,
    },
    {
        "id": 5, "name": "Netanya", "hebrew": "נתניה",
        "tier": "מטרופוליני", "tier_en": "Metropolitan", "tier_key": "metro",
        "area": "צפון", "area_en": "North",
        "position": "גלעין", "position_en": "Core",
        "passengers": 18_000,
        "modes": ["Rail", "LRT"], "total_lines": 4,
        "pop": [42_000, 32_000, 22_000], "emp": [25_000, 15_000, 8_000],
        "near_terminal": False, "terminal_type": "-",
        "terminal_weight": 0.0,
    },
]

MODE_WEIGHTS = {
    "HighSpeed Rail": 8.0, "Rail": 7.0, "Interurban Rail": 7.0,
    "Metro": 6.0, "Suburban Rail": 6.0, "LRT": 5.0, "BRT": 4.0,
    "Express Bus": 3.0, "Cable Line": 3.0, "Funicular": 2.0, "Bus": 1.0,
}

RING_WEIGHTS = [0.78, 0.15, 0.07]
REGION_WEIGHTS = {"תל אביב": 0, "ירושלים": 1, "חיפה": 1, "צפון": 1, "דרום": 1}
POSITION_WEIGHTS = {"גלעין": 3, "טבעת פנימית": 2, "טבעת חיצונית": 2, "Outer": 1}


# ===========================================================================
# Scoring calculations (match walkthrough exactly)
# ===========================================================================

def calc_activity(hub):
    return math.log10(hub["passengers"])


def calc_service_raw(hub):
    n_modes = len(hub["modes"])
    lines_per_mode = hub["total_lines"] / n_modes
    subtotal = 0.0
    mode_details = []
    for mode in hub["modes"]:
        w = MODE_WEIGHTS[mode]
        sq = math.sqrt(lines_per_mode)
        ms = w * sq
        subtotal += ms
        mode_details.append((mode, w, lines_per_mode, sq, ms))
    diversity = 1.0 + (n_modes - 1) * 0.10
    raw = subtotal * diversity
    return raw, subtotal, diversity, mode_details


def calc_location_raw(hub):
    region_w = REGION_WEIGHTS.get(hub["area"], 0.5)
    position_w = POSITION_WEIGHTS.get(hub["position"], 1.5)
    return region_w * position_w, region_w, position_w


def calc_pop_jobs_raw(hub):
    job_w = 0.80
    pop_w = 0.20
    ring_scores = []
    total = 0.0
    for i in range(3):
        weighted_val = hub["pop"][i] * pop_w + hub["emp"][i] * job_w
        rs = weighted_val * RING_WEIGHTS[i]
        ring_scores.append((hub["pop"][i], hub["emp"][i], weighted_val, RING_WEIGHTS[i], rs))
        total += rs
    return total, ring_scores


def normalize_minmax(values, vmin=None, vmax=None):
    if vmin is None:
        vmin = min(values)
    if vmax is None:
        vmax = max(values)
    if vmax == vmin:
        return [5.5] * len(values)
    return [round((v - vmin) / (vmax - vmin) * 9 + 1, 2) for v in values]


# Pre-compute all raw scores
for h in HUBS:
    h["log_pax"] = calc_activity(h)
    h["service_raw"], h["service_subtotal"], h["service_diversity"], h["service_modes"] = calc_service_raw(h)
    h["location_raw"], h["region_w"], h["position_w"] = calc_location_raw(h)
    h["popjobs_raw"], h["popjobs_rings"] = calc_pop_jobs_raw(h)
    h["terminal_raw"] = h["terminal_weight"]

# Per-tier normalization
national = [h for h in HUBS if h["tier_key"] == "national"]
metro = [h for h in HUBS if h["tier_key"] == "metro"]

# Activity (log)
nat_log = [h["log_pax"] for h in national]
met_log = [h["log_pax"] for h in metro]
nat_act_norm = normalize_minmax(nat_log)
met_act_norm = normalize_minmax(met_log)
for h, s in zip(national, nat_act_norm):
    h["activity_norm"] = s
for h, s in zip(metro, met_act_norm):
    h["activity_norm"] = s

# Service
nat_svc = [h["service_raw"] for h in national]
met_svc = [h["service_raw"] for h in metro]
nat_svc_norm = normalize_minmax(nat_svc)
met_svc_norm = normalize_minmax(met_svc)
for h, s in zip(national, nat_svc_norm):
    h["service_norm"] = s
for h, s in zip(metro, met_svc_norm):
    h["service_norm"] = s

# Location (GLOBAL)
all_loc = [h["location_raw"] for h in HUBS]
all_loc_norm = normalize_minmax(all_loc)
for h, s in zip(HUBS, all_loc_norm):
    h["location_norm"] = s

# Pop & Jobs
nat_pj = [h["popjobs_raw"] for h in national]
met_pj = [h["popjobs_raw"] for h in metro]
nat_pj_norm = normalize_minmax(nat_pj)
met_pj_norm = normalize_minmax(met_pj)
for h, s in zip(national, nat_pj_norm):
    h["popjobs_norm"] = s
for h, s in zip(metro, met_pj_norm):
    h["popjobs_norm"] = s

# Terminal (GLOBAL)
all_term = [h["terminal_raw"] for h in HUBS]
all_term_norm = normalize_minmax(all_term)
for h, s in zip(HUBS, all_term_norm):
    h["terminal_norm"] = s

# Monte Carlo approximate scores (from walkthrough)
MC_SCORES = {1: 8.15, 2: 5.78, 3: 4.55, 4: 7.55, 5: 2.85}
MC_STD = {1: 1.42, 2: 1.85, 3: 2.10, 4: 1.35, 5: 1.68}
for h in HUBS:
    h["mc_score"] = MC_SCORES[h["id"]]
    h["mc_std"] = MC_STD[h["id"]]


# ===========================================================================
# Build Excel workbook
# ===========================================================================
wb = Workbook()

# -----------------------------------------------------------------------
# SHEET 1: Hub Profiles
# -----------------------------------------------------------------------
ws1 = wb.active
ws1.title = "1. Hub Profiles"
ws1.sheet_properties.tabColor = NAVY

r = 1
write_title(ws1, r, 1, "Hub Scoring Methodology - Demo Walkthrough")
r += 1
write_title(ws1, r, 1, "5 Example Hubs: National & Metropolitan Tiers", subtitle_font)
r += 2

# -- Overview table --
write_title(ws1, r, 1, "Hub Overview", subtitle_font)
r += 1
headers = ["#", "Hub Name", "Hebrew Name", "Tier", "Area", "Metro Position", "2050 Pax/Day"]
write_header_row(ws1, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws1, r, [
        h["id"], h["name"], h["hebrew"],
        f'{h["tier"]} ({h["tier_en"]})',
        f'{h["area"]} ({h["area_en"]})',
        f'{h["position"]} ({h["position_en"]})',
        h["passengers"],
    ], fill_color=fill, bold_cols=[2], num_fmt=number_fmt_0)
    r += 1

r += 2

# -- Modes & Lines --
write_title(ws1, r, 1, "Transit Modes & Lines", subtitle_font)
r += 1
headers = ["#", "Hub", "Modes", "# Modes", "Total Lines", "Lines per Mode"]
write_header_row(ws1, r, headers)
r += 1
for h in HUBS:
    lpm = round(h["total_lines"] / len(h["modes"]), 2)
    fill = tier_fill(h["tier_key"])
    write_data_row(ws1, r, [
        h["id"], h["name"], ", ".join(h["modes"]),
        len(h["modes"]), h["total_lines"], lpm,
    ], fill_color=fill, bold_cols=[2])
    r += 1

r += 2

# -- Demographics --
write_title(ws1, r, 1, "Demographic Data (2050 Forecasts)", subtitle_font)
r += 1
headers = ["#", "Hub", "Pop 0-500m", "Pop 500-1000m", "Pop 1000-1500m",
           "Emp 0-500m", "Emp 500-1000m", "Emp 1000-1500m"]
write_header_row(ws1, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws1, r, [
        h["id"], h["name"],
        h["pop"][0], h["pop"][1], h["pop"][2],
        h["emp"][0], h["emp"][1], h["emp"][2],
    ], fill_color=fill, bold_cols=[2], num_fmt=number_fmt_0)
    r += 1

r += 2

# -- Terminal --
write_title(ws1, r, 1, "Bus Terminal Proximity", subtitle_font)
r += 1
headers = ["#", "Hub", "Near Terminal (<200m)", "Terminal Type", "Terminal Weight"]
write_header_row(ws1, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws1, r, [
        h["id"], h["name"],
        "Yes" if h["near_terminal"] else "No",
        h["terminal_type"], h["terminal_weight"],
    ], fill_color=fill, bold_cols=[2])
    r += 1

set_col_widths(ws1, {1: 5, 2: 22, 3: 22, 4: 28, 5: 22, 6: 24, 7: 18, 8: 18})

# -----------------------------------------------------------------------
# SHEET 2: Passenger Activity Score
# -----------------------------------------------------------------------
ws2 = wb.create_sheet("2. Activity Score")
ws2.sheet_properties.tabColor = DARK_BLUE

r = 1
write_title(ws2, r, 1, "Criterion 1: Passenger Activity Score")
r += 1
ws2.cell(row=r, column=1, value="Method: log₁₀(passengers) → per-tier min-max normalization to 1-10").font = section_font
r += 2

# Step 1: Log transformation
write_title(ws2, r, 1, "Step 1: Log₁₀ Transformation", subtitle_font)
r += 1
headers = ["#", "Hub", "Tier", "Passengers", "log₁₀(Passengers)"]
write_header_row(ws2, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws2, r, [
        h["id"], h["name"], h["tier"], h["passengers"], round(h["log_pax"], 3),
    ], fill_color=fill, bold_cols=[2, 5], num_fmt=number_fmt_0)
    # override the log column format
    ws2.cell(row=r, column=5).number_format = "0.000"
    r += 1

r += 2

# Step 2: Normalization
write_title(ws2, r, 1, "Step 2: Per-Tier Normalization", subtitle_font)
r += 1
ws2.cell(row=r, column=1, value="Formula: score = (value - tier_min) / (tier_max - tier_min) × 9 + 1").font = body_font
r += 2

# National tier
ws2.cell(row=r, column=1, value="National tier (ארצי)").font = section_font
nat_min = min(h["log_pax"] for h in national)
nat_max = max(h["log_pax"] for h in national)
ws2.cell(row=r, column=3, value=f"min = {nat_min:.3f}   max = {nat_max:.3f}   range = {nat_max-nat_min:.3f}").font = body_font
r += 1
headers = ["#", "Hub", "log₁₀ Value", "Tier Min", "Tier Max", "Calculation", "Normalized Score"]
write_header_row(ws2, r, headers)
r += 1
for h in national:
    ratio = (h["log_pax"] - nat_min) / (nat_max - nat_min) if nat_max != nat_min else 0
    calc_str = f'({h["log_pax"]:.3f} - {nat_min:.3f}) / {nat_max-nat_min:.3f} × 9 + 1'
    write_data_row(ws2, r, [
        h["id"], h["name"], round(h["log_pax"], 3),
        round(nat_min, 3), round(nat_max, 3),
        calc_str, h["activity_norm"],
    ], fill_color=LIGHT_BLUE, bold_cols=[2, 7])
    ws2.cell(row=r, column=3).number_format = "0.000"
    ws2.cell(row=r, column=7).number_format = "0.00"
    ws2.cell(row=r, column=7).font = score_font
    r += 1

r += 1

# Metropolitan tier
ws2.cell(row=r, column=1, value="Metropolitan tier (מטרופוליני)").font = section_font
met_min = min(h["log_pax"] for h in metro)
met_max = max(h["log_pax"] for h in metro)
ws2.cell(row=r, column=3, value=f"min = {met_min:.3f}   max = {met_max:.3f}   range = {met_max-met_min:.3f}").font = body_font
r += 1
write_header_row(ws2, r, headers)
r += 1
for h in metro:
    ratio = (h["log_pax"] - met_min) / (met_max - met_min) if met_max != met_min else 0
    calc_str = f'({h["log_pax"]:.3f} - {met_min:.3f}) / {met_max-met_min:.3f} × 9 + 1'
    write_data_row(ws2, r, [
        h["id"], h["name"], round(h["log_pax"], 3),
        round(met_min, 3), round(met_max, 3),
        calc_str, h["activity_norm"],
    ], fill_color=LIGHT_ORANGE, bold_cols=[2, 7])
    ws2.cell(row=r, column=3).number_format = "0.000"
    ws2.cell(row=r, column=7).number_format = "0.00"
    ws2.cell(row=r, column=7).font = score_font
    r += 1

set_col_widths(ws2, {1: 5, 2: 22, 3: 16, 4: 12, 5: 12, 6: 48, 7: 18})

# -----------------------------------------------------------------------
# SHEET 3: Service & Modes Score
# -----------------------------------------------------------------------
ws3 = wb.create_sheet("3. Service Score")
ws3.sheet_properties.tabColor = GREEN

r = 1
write_title(ws3, r, 1, "Criterion 2: Service & Hierarchy of Modes Score")
r += 1
ws3.cell(row=r, column=1, value="Method: Σ(mode_weight × √lines) × diversity_bonus → per-tier normalization").font = section_font
r += 2

# Mode weights reference
write_title(ws3, r, 1, "Mode Weights Reference", subtitle_font)
r += 1
headers = ["Mode", "Weight"]
write_header_row(ws3, r, headers, fill_color=GREEN)
r += 1
for mode, w in [("HighSpeed Rail", 8.0), ("Rail / Interurban", 7.0),
                ("Metro / Suburban Rail", 6.0), ("LRT", 5.0),
                ("BRT", 4.0), ("Express Bus", 3.0), ("Bus", 1.0)]:
    write_data_row(ws3, r, [mode, w], fill_color=LIGHT_GREEN)
    r += 1

r += 2

# Per-hub calculation
write_title(ws3, r, 1, "Step 1: Raw Score Calculation (per hub)", subtitle_font)
r += 1

for h in HUBS:
    fill = tier_fill(h["tier_key"])
    ws3.cell(row=r, column=1, value=f'Hub {h["id"]} - {h["name"]}').font = bold_font
    lpm = round(h["total_lines"] / len(h["modes"]), 2)
    ws3.cell(row=r, column=3, value=f'{len(h["modes"])} modes | {lpm} lines each').font = body_font
    r += 1
    headers = ["Mode", "Weight", "Lines", "√(Lines)", "Mode Score"]
    write_header_row(ws3, r, headers, fill_color=MEDIUM_BLUE)
    r += 1
    for mode, w, lcount, sq, ms in h["service_modes"]:
        write_data_row(ws3, r, [mode, w, round(lcount, 2), round(sq, 3), round(ms, 2)], fill_color=fill)
        ws3.cell(row=r, column=4).number_format = "0.000"
        ws3.cell(row=r, column=5).number_format = "0.00"
        r += 1
    # Subtotal row
    write_data_row(ws3, r, ["Subtotal", "", "", "", round(h["service_subtotal"], 2)],
                   fill_color=fill, bold_cols=[1, 5])
    ws3.cell(row=r, column=5).number_format = "0.00"
    r += 1
    # Diversity bonus
    n = len(h["modes"])
    div_text = f'{n} modes → 1 + {n-1} × 0.10 = ×{h["service_diversity"]:.2f}'
    write_data_row(ws3, r, ["Diversity Bonus", div_text, "", "", ""], fill_color=fill, bold_cols=[1, 2])
    r += 1
    # Raw score
    write_data_row(ws3, r, [
        "RAW SCORE", "", "", "",
        round(h["service_raw"], 2)
    ], fill_color=fill, bold_cols=[1, 5], font_override=score_font)
    ws3.cell(row=r, column=5).number_format = "0.00"
    ws3.cell(row=r, column=5).font = score_font
    r += 2

# Step 2: Normalization
r += 1
write_title(ws3, r, 1, "Step 2: Per-Tier Normalization", subtitle_font)
r += 1

# National
ws3.cell(row=r, column=1, value="National tier (ארצי)").font = section_font
r += 1
headers = ["#", "Hub", "Raw Score", "Tier Min", "Tier Max", "Normalized Score"]
write_header_row(ws3, r, headers)
r += 1
nat_svc_min = min(h["service_raw"] for h in national)
nat_svc_max = max(h["service_raw"] for h in national)
for h in national:
    write_data_row(ws3, r, [
        h["id"], h["name"], round(h["service_raw"], 2),
        round(nat_svc_min, 2), round(nat_svc_max, 2), h["service_norm"],
    ], fill_color=LIGHT_BLUE, bold_cols=[2, 6])
    ws3.cell(row=r, column=6).font = score_font
    ws3.cell(row=r, column=6).number_format = "0.00"
    r += 1

r += 1
# Metropolitan
ws3.cell(row=r, column=1, value="Metropolitan tier (מטרופוליני)").font = section_font
r += 1
write_header_row(ws3, r, headers)
r += 1
met_svc_min = min(h["service_raw"] for h in metro)
met_svc_max = max(h["service_raw"] for h in metro)
for h in metro:
    write_data_row(ws3, r, [
        h["id"], h["name"], round(h["service_raw"], 2),
        round(met_svc_min, 2), round(met_svc_max, 2), h["service_norm"],
    ], fill_color=LIGHT_ORANGE, bold_cols=[2, 6])
    ws3.cell(row=r, column=6).font = score_font
    ws3.cell(row=r, column=6).number_format = "0.00"
    r += 1

set_col_widths(ws3, {1: 18, 2: 22, 3: 12, 4: 12, 5: 14, 6: 18})

# -----------------------------------------------------------------------
# SHEET 4: Location Score
# -----------------------------------------------------------------------
ws4 = wb.create_sheet("4. Location Score")
ws4.sheet_properties.tabColor = ORANGE

r = 1
write_title(ws4, r, 1, "Criterion 3: Location Score (Geographic & Metropolitan)")
r += 1
ws4.cell(row=r, column=1, value="Method: region_weight × position_weight → GLOBAL normalization").font = section_font
r += 2

# Region weights reference
write_title(ws4, r, 1, "Region Weights (Periphery Prioritization)", subtitle_font)
r += 1
headers = ["Region", "Weight", "Rationale"]
write_header_row(ws4, r, headers, fill_color=ORANGE)
r += 1
for region, w, rationale in [
    ("תל אביב (Tel Aviv)", 0, "Center - already well-served"),
    ("מרכז (Center)", 0, "Center - already well-served"),
    ("ירושלים (Jerusalem)", 1, "Periphery boost"),
    ("חיפה (Haifa)", 1, "Periphery boost"),
    ("צפון (North)", 1, "Periphery boost"),
    ("דרום (South)", 1, "Periphery boost"),
]:
    write_data_row(ws4, r, [region, w, rationale], fill_color=LIGHT_ORANGE)
    r += 1

r += 1

# Position weights
write_title(ws4, r, 1, "Metropolitan Position Weights", subtitle_font)
r += 1
headers = ["Position", "Weight"]
write_header_row(ws4, r, headers, fill_color=ORANGE)
r += 1
for pos, w in [("גלעין (Core)", 3), ("טבעת (Ring)", 2), ("Outer / Periphery", 1)]:
    write_data_row(ws4, r, [pos, w], fill_color=LIGHT_ORANGE)
    r += 1

r += 2

# Step 1: Raw scores
write_title(ws4, r, 1, "Step 1: Raw Score Calculation", subtitle_font)
r += 1
headers = ["#", "Hub", "Region", "Region Wt", "Position", "Position Wt", "Raw Score"]
write_header_row(ws4, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws4, r, [
        h["id"], h["name"],
        f'{h["area"]} ({h["area_en"]})', h["region_w"],
        f'{h["position"]} ({h["position_en"]})', h["position_w"],
        h["location_raw"],
    ], fill_color=fill, bold_cols=[2, 7])
    ws4.cell(row=r, column=7).font = score_font
    r += 1

r += 2

# Step 2: Global normalization
write_title(ws4, r, 1, "Step 2: Global Normalization (all hubs together)", subtitle_font)
r += 1
loc_min = min(h["location_raw"] for h in HUBS)
loc_max = max(h["location_raw"] for h in HUBS)
ws4.cell(row=r, column=1, value=f"All 5 hubs:  min = {loc_min}   max = {loc_max}   range = {loc_max - loc_min}").font = body_font
r += 1
headers = ["#", "Hub", "Raw Score", "Calculation", "Normalized Score"]
write_header_row(ws4, r, headers)
r += 1
for h in HUBS:
    rng = loc_max - loc_min
    calc_str = f"({h['location_raw']} - {loc_min}) / {rng} × 9 + 1"
    fill = tier_fill(h["tier_key"])
    write_data_row(ws4, r, [
        h["id"], h["name"], h["location_raw"], calc_str, h["location_norm"],
    ], fill_color=fill, bold_cols=[2, 5])
    ws4.cell(row=r, column=5).font = score_font
    ws4.cell(row=r, column=5).number_format = "0.00"
    r += 1

r += 2
ws4.cell(row=r, column=1,
         value="Key: Tel Aviv hubs get minimum location score (1.0) by design — periphery prioritization for national equity.").font = Font(
    name="Calibri", italic=True, color=DARK_GRAY, size=11)

set_col_widths(ws4, {1: 5, 2: 22, 3: 22, 4: 14, 5: 26, 6: 14, 7: 18})

# -----------------------------------------------------------------------
# SHEET 5: Population & Jobs Score
# -----------------------------------------------------------------------
ws5 = wb.create_sheet("5. Pop & Jobs Score")
ws5.sheet_properties.tabColor = GOLD

r = 1
write_title(ws5, r, 1, "Criterion 4: Population & Jobs Score (2050)")
r += 1
ws5.cell(row=r, column=1, value="Method: Σ(ring_weight × (pop×pop_wt + emp×job_wt)) → per-tier normalization").font = section_font
r += 2

# Parameters
write_title(ws5, r, 1, "Distance Decay Weights (beta = 1.5)", subtitle_font)
r += 1
headers = ["Ring", "Distance", "Midpoint", "Weight"]
write_header_row(ws5, r, headers, fill_color=GOLD)
r += 1
for ring_label, dist, mid, w in [
    ("Ring 0", "0-500m", "250m", 0.78),
    ("Ring 1", "500-1000m", "750m", 0.15),
    ("Ring 2", "1000-1500m", "1250m", 0.07),
]:
    write_data_row(ws5, r, [ring_label, dist, mid, w], fill_color=LIGHT_GOLD)
    r += 1

r += 1
write_title(ws5, r, 1, "Pop/Jobs Mix: National & Metro = 80% Jobs / 20% Pop", subtitle_font)
r += 2

# Per-hub ring breakdown
write_title(ws5, r, 1, "Step 1: Raw Score Calculation (per hub)", subtitle_font)
r += 1

ring_labels = ["0-500m", "500-1000m", "1000-1500m"]
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    ws5.cell(row=r, column=1, value=f'Hub {h["id"]} - {h["name"]}').font = bold_font
    ws5.cell(row=r, column=4, value=f'{h["tier"]} — 80% jobs / 20% pop').font = body_font
    r += 1
    headers = ["Ring", "Population", "Employment", "Weighted Value", "Ring Weight", "Ring Score"]
    write_header_row(ws5, r, headers, fill_color=MEDIUM_BLUE)
    r += 1
    for i, (pop, emp, wv, rw, rs) in enumerate(h["popjobs_rings"]):
        write_data_row(ws5, r, [
            ring_labels[i], pop, emp, round(wv, 0), rw, round(rs, 0),
        ], fill_color=fill, num_fmt=number_fmt_0)
        r += 1
    # Total
    write_data_row(ws5, r, ["TOTAL", "", "", "", "", round(h["popjobs_raw"], 0)],
                   fill_color=fill, bold_cols=[1, 6], num_fmt=number_fmt_0)
    ws5.cell(row=r, column=6).font = score_font
    r += 2

# Step 2: Normalization
r += 1
write_title(ws5, r, 1, "Step 2: Per-Tier Normalization", subtitle_font)
r += 1

headers = ["#", "Hub", "Raw Score", "Tier Min", "Tier Max", "Normalized Score"]

ws5.cell(row=r, column=1, value="National tier (ארצי)").font = section_font
r += 1
write_header_row(ws5, r, headers)
r += 1
nat_pj_min = min(h["popjobs_raw"] for h in national)
nat_pj_max = max(h["popjobs_raw"] for h in national)
for h in national:
    write_data_row(ws5, r, [
        h["id"], h["name"], round(h["popjobs_raw"], 0),
        round(nat_pj_min, 0), round(nat_pj_max, 0), h["popjobs_norm"],
    ], fill_color=LIGHT_BLUE, bold_cols=[2, 6], num_fmt=number_fmt_0)
    ws5.cell(row=r, column=6).font = score_font
    ws5.cell(row=r, column=6).number_format = "0.00"
    r += 1

r += 1
ws5.cell(row=r, column=1, value="Metropolitan tier (מטרופוליני)").font = section_font
r += 1
write_header_row(ws5, r, headers)
r += 1
met_pj_min = min(h["popjobs_raw"] for h in metro)
met_pj_max = max(h["popjobs_raw"] for h in metro)
for h in metro:
    write_data_row(ws5, r, [
        h["id"], h["name"], round(h["popjobs_raw"], 0),
        round(met_pj_min, 0), round(met_pj_max, 0), h["popjobs_norm"],
    ], fill_color=LIGHT_ORANGE, bold_cols=[2, 6], num_fmt=number_fmt_0)
    ws5.cell(row=r, column=6).font = score_font
    ws5.cell(row=r, column=6).number_format = "0.00"
    r += 1

set_col_widths(ws5, {1: 14, 2: 22, 3: 14, 4: 16, 5: 14, 6: 18})

# -----------------------------------------------------------------------
# SHEET 6: Terminal Score
# -----------------------------------------------------------------------
ws6 = wb.create_sheet("6. Terminal Score")
ws6.sheet_properties.tabColor = RED

r = 1
write_title(ws6, r, 1, "Criterion 5: Bus Terminal Proximity Score")
r += 1
ws6.cell(row=r, column=1, value="Method: terminal_type_weight if within 200m, else 0 → GLOBAL normalization").font = section_font
r += 2

# Terminal type reference
write_title(ws6, r, 1, "Terminal Type Weights", subtitle_font)
r += 1
headers = ["Terminal Type", "English", "Weight"]
write_header_row(ws6, r, headers, fill_color=RED)
r += 1
for heb, eng, w in [
    ("מתקן משולב", "Integrated facility", 3.0),
    ("מסוף גדול", "Large terminal", 3.0),
    ("מסוף בינוני", "Medium terminal", 2.0),
    ("מסוף קטן", "Small terminal", 2.0),
    ("חניון לילה", "Night parking", 1.0),
]:
    write_data_row(ws6, r, [heb, eng, w], fill_color=LIGHT_RED)
    r += 1

r += 2

# Step 1: Raw
write_title(ws6, r, 1, "Step 1: Raw Score", subtitle_font)
r += 1
headers = ["#", "Hub", "Near Terminal?", "Terminal Type", "Raw Score"]
write_header_row(ws6, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws6, r, [
        h["id"], h["name"],
        "Yes" if h["near_terminal"] else "No",
        h["terminal_type"], h["terminal_raw"],
    ], fill_color=fill, bold_cols=[2, 5])
    ws6.cell(row=r, column=5).font = score_font
    r += 1

r += 2

# Step 2: Global normalization
write_title(ws6, r, 1, "Step 2: Global Normalization", subtitle_font)
r += 1
term_min = min(h["terminal_raw"] for h in HUBS)
term_max = max(h["terminal_raw"] for h in HUBS)
ws6.cell(row=r, column=1, value=f"All 5 hubs:  min = {term_min}   max = {term_max}   range = {term_max - term_min}").font = body_font
r += 1
headers = ["#", "Hub", "Raw Score", "Calculation", "Normalized Score"]
write_header_row(ws6, r, headers)
r += 1
for h in HUBS:
    rng = term_max - term_min
    calc_str = f"({h['terminal_raw']:.1f} - {term_min:.1f}) / {rng:.1f} × 9 + 1"
    fill = tier_fill(h["tier_key"])
    write_data_row(ws6, r, [
        h["id"], h["name"], h["terminal_raw"], calc_str, h["terminal_norm"],
    ], fill_color=fill, bold_cols=[2, 5])
    ws6.cell(row=r, column=5).font = score_font
    ws6.cell(row=r, column=5).number_format = "0.00"
    r += 1

set_col_widths(ws6, {1: 5, 2: 22, 3: 20, 4: 30, 5: 18})

# -----------------------------------------------------------------------
# SHEET 7: Score Summary
# -----------------------------------------------------------------------
ws7 = wb.create_sheet("7. Score Summary")
ws7.sheet_properties.tabColor = NAVY

r = 1
write_title(ws7, r, 1, "Score Summary - All 5 Criteria")
r += 2

# Before normalization
write_title(ws7, r, 1, "Before Normalization (Raw Scores)", subtitle_font)
r += 1
headers = ["#", "Hub", "Tier", "Activity (pax)", "Service (raw)", "Location (raw)",
           "Pop & Jobs (raw)", "Terminal (raw)"]
write_header_row(ws7, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws7, r, [
        h["id"], h["name"], h["tier"],
        h["passengers"], round(h["service_raw"], 2),
        h["location_raw"], round(h["popjobs_raw"], 0), h["terminal_raw"],
    ], fill_color=fill, bold_cols=[2], num_fmt=number_fmt_0)
    ws7.cell(row=r, column=5).number_format = "0.00"
    r += 1

r += 2

# After normalization
write_title(ws7, r, 1, "After Normalization (1-10 Scale)", subtitle_font)
r += 1
headers = ["#", "Hub", "Tier", "Activity", "Service", "Location",
           "Pop & Jobs", "Terminal", "Simple Mean"]
write_header_row(ws7, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    scores = [h["activity_norm"], h["service_norm"], h["location_norm"],
              h["popjobs_norm"], h["terminal_norm"]]
    mean_score = round(sum(scores) / len(scores), 2)
    write_data_row(ws7, r, [
        h["id"], h["name"], h["tier"],
        h["activity_norm"], h["service_norm"], h["location_norm"],
        h["popjobs_norm"], h["terminal_norm"], mean_score,
    ], fill_color=fill, bold_cols=[2, 9])
    for c in range(4, 10):
        ws7.cell(row=r, column=c).number_format = "0.00"
    ws7.cell(row=r, column=9).font = score_font
    r += 1

r += 2

# Normalization method summary
write_title(ws7, r, 1, "Normalization Method Summary", subtitle_font)
r += 1
headers = ["Criterion", "Normalization", "Rationale"]
write_header_row(ws7, r, headers)
r += 1
for crit, norm, rationale in [
    ("Passenger Activity", "Per Tier + log₁₀", "Fair comparison within tier; log prevents mega-station dominance"),
    ("Service & Modes", "Per Tier", "Fair comparison within tier"),
    ("Location", "GLOBAL", "Consistent geographic equity signal across all tiers"),
    ("Population & Jobs", "Per Tier", "Fair comparison within tier"),
    ("Bus Terminal", "GLOBAL", "Consistent terminal integration signal across all tiers"),
]:
    write_data_row(ws7, r, [crit, norm, rationale], bold_cols=[2])
    r += 1

set_col_widths(ws7, {1: 5, 2: 22, 3: 16, 4: 14, 5: 14, 6: 14, 7: 16, 8: 14, 9: 14})

# -----------------------------------------------------------------------
# SHEET 8: Monte Carlo
# -----------------------------------------------------------------------
ws8 = wb.create_sheet("8. Monte Carlo")
ws8.sheet_properties.tabColor = DARK_BLUE

r = 1
write_title(ws8, r, 1, "Monte Carlo Aggregation (10,000 Iterations)")
r += 2

# How it works
write_title(ws8, r, 1, "How It Works", subtitle_font)
r += 1
steps = [
    "1. Generate 5 random weights (each between 0% and 50%)",
    "2. Normalize weights to sum to 100%",
    "3. Calculate weighted score for each hub",
    "4. Repeat 10,000 times",
    "5. Final score = average across all iterations",
]
for step in steps:
    ws8.cell(row=r, column=1, value=step).font = body_font
    r += 1

r += 1

# Example iteration
write_title(ws8, r, 1, "Example: One Iteration", subtitle_font)
r += 1
headers = ["Criterion", "Random Draw", "Normalized Weight"]
write_header_row(ws8, r, headers, fill_color=MEDIUM_BLUE)
r += 1
example_draws = [("Activity", 0.35), ("Service", 0.28), ("Location", 0.42),
                 ("Pop & Jobs", 0.15), ("Terminal", 0.25)]
total_draw = sum(d for _, d in example_draws)
for crit, draw in example_draws:
    norm_w = draw / total_draw
    write_data_row(ws8, r, [crit, draw, norm_w], fill_color=LIGHT_BLUE)
    ws8.cell(row=r, column=3).number_format = "0.0%"
    r += 1
write_data_row(ws8, r, ["SUM", total_draw, 1.0], fill_color=LIGHT_BLUE, bold_cols=[1, 2, 3])
ws8.cell(row=r, column=3).number_format = "0.0%"
r += 2

# Hub 1 example calc
ws8.cell(row=r, column=1, value="Hub 1 score for this iteration:").font = bold_font
r += 1
ws8.cell(row=r, column=1,
         value="10.00×0.241 + 10.00×0.193 + 1.00×0.290 + 10.00×0.103 + 10.00×0.172 = 7.38").font = body_font
r += 2

# MC Results
write_title(ws8, r, 1, "Monte Carlo Results (10,000 iterations, seed=42)", subtitle_font)
r += 1
headers = ["#", "Hub", "Tier", "Activity", "Service", "Location", "Pop/Jobs",
           "Terminal", "MC Final Score", "Std Dev"]
write_header_row(ws8, r, headers)
r += 1
for h in HUBS:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws8, r, [
        h["id"], h["name"], h["tier"],
        h["activity_norm"], h["service_norm"], h["location_norm"],
        h["popjobs_norm"], h["terminal_norm"],
        h["mc_score"], h["mc_std"],
    ], fill_color=fill, bold_cols=[2, 9])
    for c in range(4, 11):
        ws8.cell(row=r, column=c).number_format = "0.00"
    ws8.cell(row=r, column=9).font = score_font
    r += 1

r += 2
ws8.cell(row=r, column=1,
         value="Std Dev shows sensitivity to weight changes. Higher = more sensitive to how criteria are weighted.").font = Font(
    name="Calibri", italic=True, color=DARK_GRAY, size=11)

set_col_widths(ws8, {1: 5, 2: 22, 3: 16, 4: 12, 5: 12, 6: 12, 7: 12, 8: 12, 9: 16, 10: 12})

# -----------------------------------------------------------------------
# SHEET 9: Final Ranking
# -----------------------------------------------------------------------
ws9 = wb.create_sheet("9. Final Ranking")
ws9.sheet_properties.tabColor = NAVY

r = 1
write_title(ws9, r, 1, "Final Ranking")
r += 2

# Ranking rules
write_title(ws9, r, 1, "Ranking Rules", subtitle_font)
r += 1
rules = [
    "National (ארצי): All national hubs ranked GLOBALLY",
    "Metropolitan (מטרופוליני): Ranked WITHIN their geographic area",
    "Local (עירוני): Ranked WITHIN their geographic area",
]
for rule in rules:
    ws9.cell(row=r, column=1, value=rule).font = body_font
    r += 1

r += 1

# National ranking
write_title(ws9, r, 1, "National Ranking (Global)", subtitle_font)
r += 1
headers = ["Rank", "Hub", "MC Score", "Key Strengths", "Key Weaknesses"]
write_header_row(ws9, r, headers, fill_color=DARK_BLUE)
r += 1
national_sorted = sorted(national, key=lambda h: h["mc_score"], reverse=True)
strengths = {
    1: "Highest activity, most modes, most jobs",
    2: "Strong location, large terminal",
    3: "Strong location, large terminal",
}
weaknesses = {
    1: "Low location score (central area)",
    2: "Lower service diversity",
    3: "Lowest activity among national hubs",
}
for rank, h in enumerate(national_sorted, 1):
    write_data_row(ws9, r, [
        rank, h["name"], h["mc_score"],
        strengths[h["id"]], weaknesses[h["id"]],
    ], fill_color=LIGHT_BLUE, bold_cols=[1, 2, 3])
    ws9.cell(row=r, column=3).number_format = "0.00"
    ws9.cell(row=r, column=3).font = score_font
    r += 1

r += 2

# Metropolitan ranking by area
write_title(ws9, r, 1, "Metropolitan Ranking (Per Area)", subtitle_font)
r += 1
for area_heb, area_en in [("תל אביב", "Tel Aviv"), ("צפון", "North")]:
    ws9.cell(row=r, column=1, value=f"{area_heb} ({area_en}) Area").font = section_font
    r += 1
    headers = ["Rank", "Hub", "MC Score"]
    write_header_row(ws9, r, headers, fill_color=ORANGE)
    r += 1
    area_hubs = [h for h in metro if h["area"] == area_heb]
    area_hubs.sort(key=lambda h: h["mc_score"], reverse=True)
    for rank, h in enumerate(area_hubs, 1):
        write_data_row(ws9, r, [rank, h["name"], h["mc_score"]],
                       fill_color=LIGHT_ORANGE, bold_cols=[1, 2, 3])
        ws9.cell(row=r, column=3).number_format = "0.00"
        ws9.cell(row=r, column=3).font = score_font
        r += 1
    r += 1

r += 1

# Complete results table
write_title(ws9, r, 1, "Complete Results Table", subtitle_font)
r += 1
headers = ["Rank (Group)", "Hub", "Tier", "Area", "Activity", "Service",
           "Location", "Pop/Jobs", "Terminal", "MC Score"]
write_header_row(ws9, r, headers)
r += 1

# Build ordered results
result_rows = []
for rank, h in enumerate(national_sorted, 1):
    result_rows.append((f"National #{rank}", h))
for area_heb in ["תל אביב", "צפון"]:
    area_hubs = sorted([h for h in metro if h["area"] == area_heb],
                       key=lambda x: x["mc_score"], reverse=True)
    for rank, h in enumerate(area_hubs, 1):
        area_short = "TA" if area_heb == "תל אביב" else "North"
        result_rows.append((f"Metro {area_short} #{rank}", h))

for rank_label, h in result_rows:
    fill = tier_fill(h["tier_key"])
    write_data_row(ws9, r, [
        rank_label, h["name"], h["tier"], h["area"],
        h["activity_norm"], h["service_norm"], h["location_norm"],
        h["popjobs_norm"], h["terminal_norm"], h["mc_score"],
    ], fill_color=fill, bold_cols=[1, 2, 10])
    for c in range(5, 11):
        ws9.cell(row=r, column=c).number_format = "0.00"
    ws9.cell(row=r, column=10).font = score_font
    r += 1

r += 2

# Key takeaways
write_title(ws9, r, 1, "Key Takeaways", subtitle_font)
r += 1
takeaways = [
    "1. Log transformation: prevents mega-stations from dominating (150K vs 60K → 1.08x ratio, not 2.5x)",
    "2. Per-tier normalization: hubs compete within their own tier (32K pax = 10/10 among metro hubs)",
    "3. Location equity: Tel Aviv hubs get min location score (1.0) — periphery boosted to 10.0",
    "4. Diversity bonus: 4 modes = +30% bonus vs 2 modes = +10% (rewards multimodality)",
    "5. Monte Carlo: 10,000 random weight sets prevent any single criterion from dominating",
    "6. Area-based ranking: metro hubs ranked within area, not compared to different regions",
]
for t in takeaways:
    ws9.cell(row=r, column=1, value=t).font = body_font
    r += 1

set_col_widths(ws9, {1: 18, 2: 22, 3: 16, 4: 14, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 14})

# ---------------------------------------------------------------------------
# Freeze panes and print settings for all sheets
# ---------------------------------------------------------------------------
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
wb.save(OUTPUT_PATH)
print(f"Demo Excel saved to: {OUTPUT_PATH}")
